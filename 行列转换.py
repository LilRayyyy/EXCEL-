import sys
import pandas as pd
import numpy as np
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QFileDialog, QLabel, QTextEdit
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

class TableConverterApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Excel Table Converter')
        self.setGeometry(300, 300, 400, 300)

        layout = QVBoxLayout()

        self.file_label = QLabel('No file selected', self)
        layout.addWidget(self.file_label)

        select_button = QPushButton('Select File', self)
        select_button.clicked.connect(self.select_file)
        layout.addWidget(select_button)

        convert_button = QPushButton('Convert', self)
        convert_button.clicked.connect(self.convert_table)
        layout.addWidget(convert_button)

        self.log_text = QTextEdit(self)
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)

        self.setLayout(layout)

        self.filename = None

    def select_file(self):
        self.filename, _ = QFileDialog.getOpenFileName(self, "Select Excel file", "", "Excel Files (*.xlsx)")
        if self.filename:
            self.file_label.setText(f'Selected: {self.filename}')

    def log(self, message):
        self.log_text.append(message)

    def convert_table(self):
        if not self.filename:
            self.log('Please select a file first')
            return

        self.log('Starting conversion...')

        # Read the Excel file
        df = pd.read_excel(self.filename)
        self.log(f'Read {len(df)} rows from the input file')

        # Ensure 'Models' column exists
        if 'Models' not in df.columns:
            self.log('Error: "Models" column not found in the file')
            return

        # List of methods we're interested in
        methods = ['Voting', 'Averaging', 'Bagging', 'Stacking', 'AdaBoost']

        # Get all unique model combinations
        all_models = df['Models'].unique()
        self.log(f'Found {len(all_models)} unique model combinations')

        # Create a new DataFrame with all possible combinations
        result_df = pd.DataFrame({'Models': all_models})

        # Process each method
        for method in methods:
            method_df = df[df['Method'] == method].copy()
            method_df = method_df.set_index('Models')['R2 Score']
            result_df = result_df.merge(method_df.rename(method), on='Models', how='left')

        self.log(f'Processed data: {len(result_df)} rows, {len(result_df.columns)} columns')

        # Replace NaN with empty string
        result_df = result_df.replace(np.nan, '', regex=True)

        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active

        # Write the header
        ws.append(['Method'] + methods)

        # Write the data
        for r in dataframe_to_rows(result_df, index=False, header=False):
            ws.append(r)

        # Style the header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Style the index column
        for cell in ws['A']:
            cell.alignment = Alignment(horizontal='left')

        # Save the workbook
        output_filename = self.filename.replace('.xlsx', '_converted.xlsx')
        wb.save(output_filename)

        self.log(f'Converted file saved as: {output_filename}')
        self.log(f'Conversion complete. Output file has {ws.max_row} rows.')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = TableConverterApp()
    ex.show()
    sys.exit(app.exec_())